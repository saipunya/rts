from __future__ import annotations

import sys
from pathlib import Path


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python3 scripts/inspect_xlsx.py <file.xlsx>")
        return 2

    path = Path(sys.argv[1])
    print("file:", path)
    print("exists:", path.exists())
    if not path.exists():
        return 2
    print("size:", path.stat().st_size)

    try:
        import openpyxl  # type: ignore
    except Exception as e:
        print("ERROR: openpyxl not available:", repr(e))
        print("Try: python3 -m pip install --user openpyxl")
        return 3

    print("openpyxl:", getattr(openpyxl, "__version__", "?"))

    wb = openpyxl.load_workbook(path, data_only=True)
    print("sheets:", wb.sheetnames)

    for name in wb.sheetnames:
        ws = wb[name]

        # Find first non-empty row (scan first 200 rows)
        first_row = None
        scan_max_row = min(ws.max_row or 1, 200)
        scan_max_col = min(ws.max_column or 1, 80)
        for r in range(1, scan_max_row + 1):
            if any((ws.cell(r, c).value not in (None, "")) for c in range(1, scan_max_col + 1)):
                first_row = r
                break

        print("\n==", name, "==")
        print("max_row:", ws.max_row, "max_col:", ws.max_column, "first_nonempty_row:", first_row)
        if first_row is None:
            continue

        end_row = min(ws.max_row or first_row, first_row + 10)
        end_col = min(ws.max_column or 1, 25)
        for r in range(first_row, end_row + 1):
            row_vals = []
            for c in range(1, end_col + 1):
                v = ws.cell(r, c).value
                if isinstance(v, str):
                    v = v.strip()
                row_vals.append(v)
            print(f"R{r}:", row_vals)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
